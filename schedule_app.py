import sys
import json
import os
import socket
import threading
import base64
import tempfile
import getpass
from collections import defaultdict
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTableWidget, QTableWidgetItem,
                             QPushButton, QVBoxLayout, QWidget, QHBoxLayout, QMenu,
                             QLabel, QMessageBox, QFileDialog, QHeaderView,
                             QAction, QComboBox, QInputDialog, QDialog, 
                             QVBoxLayout, QCheckBox, QScrollArea, QDialogButtonBox,
                             QShortcut, QTimeEdit, QFormLayout, QGridLayout, QSplitter,
                             QLineEdit, QCheckBox, QMenuBar, QListWidget)
from PyQt5.QtCore import Qt, QSettings, QThread, pyqtSignal, QTime, QDate
from PyQt5.QtGui import QColor, QKeySequence, QFont, QPainter, QIcon, QIntValidator
from datetime import datetime
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


class ServerConnectionDialog(QDialog):
    def __init__(self, parent=None, current_host='localhost', current_port=8888, current_network_mode=False):
        super().__init__(parent)
        self.setWindowTitle("Настройки подключения к серверу")
        self.setModal(True)
        self.setFixedSize(350, 200)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        form_layout = QFormLayout()
        
        self.host_edit = QLineEdit(current_host)
        self.port_edit = QLineEdit(str(current_port))
        self.port_edit.setValidator(QIntValidator(1, 65535))
        
        form_layout.addRow("Хост сервера:", self.host_edit)
        form_layout.addRow("Порт:", self.port_edit)
        
        layout.addLayout(form_layout)
        
        self.use_network_cb = QCheckBox("Использовать сетевой режим")
        self.use_network_cb.setChecked(current_network_mode)
        layout.addWidget(self.use_network_cb)
        
        # Информация о текущем подключении
        self.status_label = QLabel()
        self.update_status_label(current_network_mode, current_host, current_port)
        layout.addWidget(self.status_label)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def update_status_label(self, network_mode, host, port):
        if network_mode:
            self.status_label.setText(f"Текущий режим: СЕТЕВОЙ\nСервер: {host}:{port}")
            self.status_label.setStyleSheet("color: green; font-weight: bold;")
        else:
            self.status_label.setText("Текущий режим: ЛОКАЛЬНЫЙ")
            self.status_label.setStyleSheet("color: blue; font-weight: bold;")
    
    def get_connection_params(self):
        return {
            'use_network': self.use_network_cb.isChecked(),
            'host': self.host_edit.text(),
            'port': int(self.port_edit.text()) if self.port_edit.text() else 8888
        }


class NetworkManager:
    def __init__(self):
        self.socket = None
        self.host = 'localhost'
        self.port = 8888
        self.is_connected = False
        
    def connect_to_server(self, host=None, port=None):
        """Подключение к серверу"""
        if host:
            self.host = host
        if port:
            self.port = port
            
        try:
            self.socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            self.socket.settimeout(5)
            self.socket.connect((self.host, self.port))
            self.is_connected = True
            return True
        except Exception as e:
            self.is_connected = False
            return False
    
    def disconnect(self):
        """Отключение от сервера"""
        self.is_connected = False
        if self.socket:
            self.socket.close()
            self.socket = None
    
    def send_request(self, request):
        """Отправка запроса на сервер"""
        if not self.is_connected or not self.socket:
            return None
            
        try:
            data = json.dumps(request).encode('utf-8')
            self.socket.send(data)
            
            # Ждем ответ
            response_data = self.socket.recv(10240).decode('utf-8')
            if response_data:
                return json.loads(response_data)
            return None
        except Exception as e:
            return None


class SharedFilesDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Общие Excel файлы")
        self.setModal(True)
        self.setMinimumSize(600, 500)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Заголовок
        layout.addWidget(QLabel("Доступные Excel файлы на сервере:"))
        
        # Таблица файлов
        self.files_table = QTableWidget()
        self.files_table.setColumnCount(5)
        self.files_table.setHorizontalHeaderLabels(["ID", "Имя файла", "Размер", "Автор", "Дата создания"])
        self.files_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.files_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.files_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.files_table)
        
        # Кнопки
        button_layout = QHBoxLayout()
        
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.clicked.connect(self.refresh_files)
        button_layout.addWidget(self.refresh_btn)
        
        self.download_btn = QPushButton("Скачать")
        self.download_btn.clicked.connect(self.download_file)
        button_layout.addWidget(self.download_btn)
        
        self.open_btn = QPushButton("Открыть")
        self.open_btn.clicked.connect(self.open_file)
        button_layout.addWidget(self.open_btn)
        
        self.delete_btn = QPushButton("Удалить")
        self.delete_btn.clicked.connect(self.delete_file)
        button_layout.addWidget(self.delete_btn)
        
        layout.addLayout(button_layout)
        
        # Статус
        self.status_label = QLabel("")
        layout.addWidget(self.status_label)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Close)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
        self.refresh_files()
    
    def refresh_files(self):
        """Обновить список файлов"""
        self.files_table.setRowCount(0)
        files = self.db.get_available_excel_files()
        
        if not files:
            self.status_label.setText("Нет доступных файлов на сервере")
            return
        
        self.files_table.setRowCount(len(files))
        
        for row, file_info in enumerate(files):
            # ID
            id_item = QTableWidgetItem(str(file_info['id']))
            id_item.setData(Qt.UserRole, file_info)
            self.files_table.setItem(row, 0, id_item)
            
            # Имя файла
            self.files_table.setItem(row, 1, QTableWidgetItem(file_info['file_name']))
            
            # Размер
            size_kb = file_info['file_size'] / 1024
            self.files_table.setItem(row, 2, QTableWidgetItem(f"{size_kb:.1f} KB"))
            
            # Автор
            self.files_table.setItem(row, 3, QTableWidgetItem(file_info['created_by']))
            
            # Дата
            created_at = file_info['created_at']
            if 'T' in created_at:
                created_at = created_at.split('T')[0]
            self.files_table.setItem(row, 4, QTableWidgetItem(created_at))
        
        self.status_label.setText(f"Найдено файлов: {len(files)}")
    
    def get_selected_file(self):
        """Получить выбранный файл"""
        current_row = self.files_table.currentRow()
        if current_row >= 0:
            item = self.files_table.item(current_row, 0)
            return item.data(Qt.UserRole)
        return None
    
    def download_file(self):
        """Скачать выбранный файл"""
        file_info = self.get_selected_file()
        if not file_info:
            QMessageBox.warning(self, "Ошибка", "Выберите файл для скачивания")
            return
        
        file_name = file_info['file_name']
        file_id = file_info['id']
        
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        if folder:
            save_path = os.path.join(folder, file_name)
            if self.db.get_excel_from_server(file_name, save_path, file_id):
                QMessageBox.information(self, "Успех", f"Файл скачан:\n{save_path}")
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось скачать файл")
    
    def open_file(self):
        """Открыть выбранный файл"""
        file_info = self.get_selected_file()
        if not file_info:
            QMessageBox.warning(self, "Ошибка", "Выберите файл для открытия")
            return
        
        file_name = file_info['file_name']
        file_id = file_info['id']
        
        # Создаем временный файл
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, file_name)
        
        if self.db.get_excel_from_server(file_name, temp_path, file_id):
            try:
                if sys.platform == "win32":
                    os.startfile(temp_path)
                elif sys.platform == "darwin":
                    os.system(f'open "{temp_path}"')
                else:
                    os.system(f'xdg-open "{temp_path}"')
                self.status_label.setText(f"Файл открыт: {file_name}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось открыть файл: {str(e)}")
        else:
            QMessageBox.warning(self, "Ошибка", "Не удалось загрузить файл с сервера")
    
    def delete_file(self):
        """Удалить выбранный файл"""
        file_info = self.get_selected_file()
        if not file_info:
            QMessageBox.warning(self, "Ошибка", "Выберите файл для удаления")
            return
        
        file_name = file_info['file_name']
        file_id = file_info['id']
        
        reply = QMessageBox.question(
            self, 
            "Подтверждение удаления",
            f"Вы уверены, что хотите удалить файл '{file_name}'?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.db.delete_excel_file(file_id):
                QMessageBox.information(self, "Успех", "Файл удален")
                self.refresh_files()
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось удалить файл")


class MonthSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent) 
        self.setWindowTitle("Выбор месяца и года")
        self.setModal(True)
        self.setMinimumWidth(300)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        grid_layout = QGridLayout()
        
        # Текущая дата для значений по умолчанию
        current_date = QDate.currentDate()
        
        # Год
        grid_layout.addWidget(QLabel("Год:"), 0, 0)
        self.year_combo = QComboBox()
        current_year = current_date.year()
        for year in range(current_year - 5, current_year + 6):  # 5 лет назад - 5 лет вперед
            self.year_combo.addItem(str(year), year)
        self.year_combo.setCurrentText(str(current_year))
        grid_layout.addWidget(self.year_combo, 0, 1)
        
        # Месяц
        grid_layout.addWidget(QLabel("Месяц:"), 1, 0)
        self.month_combo = QComboBox()
        month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
        }
        for month_num, month_name in month_names.items():
            self.month_combo.addItem(month_name, month_num)
        self.month_combo.setCurrentIndex(current_date.month() - 1)
        grid_layout.addWidget(self.month_combo, 1, 1)
        
        layout.addLayout(grid_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_selected_period(self):
        year = self.year_combo.currentData()
        month = self.month_combo.currentData()
        return f"{year}-{month:02d}"


class NoteDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Примечание к рабочему дню")
        self.setModal(True)
        self.setMinimumWidth(300)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        form_layout = QFormLayout()
        
        self.end_time_edit = QTimeEdit()
        self.end_time_edit.setTime(QTime(20, 0))
        self.end_time_edit.setDisplayFormat("HH:mm")
        
        form_layout.addRow("Время окончания:", self.end_time_edit)
        
        layout.addLayout(form_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def get_end_time(self):
        return self.end_time_edit.time().toString("HH:mm")


class NoteTableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.is_loading = False  # Флаг для отслеживания процесса загрузки
    
    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self.viewport())
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor(255, 0, 0))
        painter.end()
    
    def setItem(self, row, column, item):
        super().setItem(row, column, item)
        # Обновляем счетчики после изменения ячейки, но не во время загрузки
        if not self.is_loading and column < self.columnCount() - 4:
            self.parent.update_counters()


class SaveThread(QThread):
    finished = pyqtSignal(bool, str)
    
    def __init__(self, db, period, data):
        super().__init__()
        self.db = db
        self.period = period
        self.data = data
    
    def run(self):
        try:
            self.db.save_schedule(self.period, self.data)
            self.finished.emit(True, self.period)
        except Exception as e:
            self.finished.emit(False, str(e))


class EmployeeSelectionDialog(QDialog):
    def __init__(self, employees, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор сотрудников")
        self.setModal(True)
        self.setMinimumWidth(350)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        if not employees:
            layout.addWidget(QLabel("Нет доступных сотрудников"))
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(self.reject)
            layout.addWidget(button_box)
            return
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.content = QWidget()
        self.scroll_layout = QVBoxLayout(self.content)
        
        self.checkboxes = []
        for emp in employees:
            cb = QCheckBox(emp["name"])
            cb.setChecked(True)
            self.checkboxes.append((emp["name"], cb))
            self.scroll_layout.addWidget(cb)
        
        scroll.setWidget(self.content)
        layout.addWidget(scroll)
        
        btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("Выбрать все")
        self.select_all_btn.clicked.connect(lambda: self.set_all_checkboxes(True))
        btn_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = QPushButton("Снять все")
        self.deselect_all_btn.clicked.connect(lambda: self.set_all_checkboxes(False))
        btn_layout.addWidget(self.deselect_all_btn)
        
        layout.addLayout(btn_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def set_all_checkboxes(self, state):
        for _, cb in self.checkboxes:
            cb.setChecked(state)
    
    def get_selected_employees(self):
        return [name for name, cb in self.checkboxes if cb.isChecked()]


class AddEmployeeToPeriodDialog(QDialog):
    def __init__(self, all_employees, current_employees, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить сотрудников в месяц")
        self.setModal(True)
        self.setMinimumWidth(350)
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Фильтруем сотрудников: только те, кого еще нет в текущем периоде
        current_employee_names = [emp["name"] for emp in current_employees]
        available_employees = [emp for emp in all_employees if emp["name"] not in current_employee_names]
        
        if not available_employees:
            layout.addWidget(QLabel("Нет доступных сотрудников для добавления"))
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(self.reject)
            layout.addWidget(button_box)
            return
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.content = QWidget()
        self.scroll_layout = QVBoxLayout(self.content)
        
        self.checkboxes = []
        for emp in available_employees:
            cb = QCheckBox(emp["name"])
            cb.setChecked(True)
            self.checkboxes.append((emp["name"], cb))
            self.scroll_layout.addWidget(cb)
        
        scroll.setWidget(self.content)
        layout.addWidget(scroll)
        
        btn_layout = QHBoxLayout()
        self.select_all_btn = QPushButton("Выбрать все")
        self.select_all_btn.clicked.connect(lambda: self.set_all_checkboxes(True))
        btn_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = QPushButton("Снять все")
        self.deselect_all_btn.clicked.connect(lambda: self.set_all_checkboxes(False))
        btn_layout.addWidget(self.deselect_all_btn)
        
        layout.addLayout(btn_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def set_all_checkboxes(self, state):
        for _, cb in self.checkboxes:
            cb.setChecked(state)
    
    def get_selected_employees(self):
        return [name for name, cb in self.checkboxes if cb.isChecked()]


class NoteItem(QTableWidgetItem):
    def __init__(self, text="", has_note=False, note_data=None):
        super().__init__(text)
        self.has_note = has_note
        self.note_data = note_data or {}
    
    def clone(self):
        return NoteItem(self.text(), self.has_note, self.note_data.copy())


class ScheduleManager:
    def __init__(self, use_network=False, server_host='localhost', server_port=8888):
        self.use_network = use_network
        self.network_manager = None
        self.server_host = server_host
        self.server_port = server_port
        
        if use_network:
            self.network_manager = NetworkManager()
            if not self.network_manager.connect_to_server(server_host, server_port):
                raise Exception("Не удалось подключиться к серверу")
        else:
            # Локальный режим (оригинальный код)
            self.schedule_folder = "schedules"
            self.employees_file = "employees.json"
            os.makedirs(self.schedule_folder, exist_ok=True)
            self.employees = self.load_employees()
            self._cache = {}
    
    def reconnect(self, use_network=None, server_host=None, server_port=None):
        """Переподключение с новыми параметрами"""
        if use_network is not None:
            self.use_network = use_network
        if server_host is not None:
            self.server_host = server_host
        if server_port is not None:
            self.server_port = server_port
            
        if self.network_manager:
            self.network_manager.disconnect()
            
        if self.use_network:
            self.network_manager = NetworkManager()
            return self.network_manager.connect_to_server(self.server_host, self.server_port)
        else:
            return True
    
    def load_employees(self):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({'action': 'get_employees'})
            if response and response.get('status') == 'success':
                return response.get('data', [])
            return []
        else:
            # Локальный режим
            if os.path.exists(self.employees_file):
                try:
                    with open(self.employees_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
                except:
                    return []
            return []
    
    def save_employees(self):
        if self.use_network:
            return True  # В сетевом режиме сотрудники сохраняются на сервере
        else:
            try:
                with open(self.employees_file, 'w', encoding='utf-8') as f:
                    json.dump(self.employees, f, ensure_ascii=False, indent=2)
                return True
            except:
                return False
    
    def get_periods(self):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({'action': 'get_periods'})
            if response and response.get('status') == 'success':
                return [item['period'] for item in response.get('data', [])]
            return []
        else:
            # Локальный режим
            periods = []
            for filename in os.listdir(self.schedule_folder):
                if filename.endswith(".json"):
                    periods.append(filename[:-5])
            return sorted(periods)
    
    def load_schedule(self, period):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({
                'action': 'get_schedule',
                'period': period
            })
            if response and response.get('status') == 'success':
                return response.get('data')
            return None
        else:
            # Локальный режим
            if period in self._cache:
                return self._cache[period]
                
            filepath = os.path.join(self.schedule_folder, f"{period}.json")
            if os.path.exists(filepath):
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        if 'notes' not in data:
                            data['notes'] = {}
                        self._cache[period] = data
                        return data
                except:
                    return None
            return None
    
    def save_schedule(self, period, data):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({
                'action': 'save_schedule',
                'period': period,
                'schedule_data': data
            })
            return response and response.get('status') == 'success'
        else:
            # Локальный режим
            filepath = os.path.join(self.schedule_folder, f"{period}.json")
            try:
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                self._cache[period] = data
                return True
            except:
                return False
    
    def add_employee(self, name):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({
                'action': 'add_employee',
                'name': name
            })
            return response and response.get('status') == 'success'
        else:
            # Локальный режим
            if not any(emp["name"].lower() == name.lower() for emp in self.employees):
                self.employees.append({"name": name, "position": ""})
                return self.save_employees()
            return False
    
    def create_period(self, period, employee_names):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            # Получаем ID сотрудников
            employees = self.load_employees()
            employee_ids = []
            for emp in employees:
                if emp['name'] in employee_names:
                    employee_ids.append(emp['id'])
            
            response = self.network_manager.send_request({
                'action': 'create_period',
                'period': period,
                'employee_ids': employee_ids
            })
            return response and response.get('status') == 'success'
        else:
            # Локальный режим
            try:
                year, month = map(int, period.split('-'))
                days_in_month = monthrange(year, month)[1]
            except:
                days_in_month = 31
            
            selected_employees = [emp for emp in self.employees if emp["name"] in employee_names]
            
            return self.save_schedule(period, {
                "employees": selected_employees,
                "schedule": {emp["name"]: [4]*days_in_month for emp in selected_employees},
                "notes": {}
            })
    
    def add_employees_to_period(self, period, employee_names):
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            employees = self.load_employees()
            employee_ids = []
            for emp in employees:
                if emp['name'] in employee_names:
                    employee_ids.append(emp['id'])
            
            response = self.network_manager.send_request({
                'action': 'add_employees_to_period',
                'period': period,
                'employee_ids': employee_ids
            })
            return response and response.get('status') == 'success'
        else:
            # Локальный режим
            schedule_data = self.load_schedule(period)
            if not schedule_data:
                return False
            
            new_employees = [emp for emp in self.employees if emp["name"] in employee_names and emp not in schedule_data["employees"]]
            
            try:
                year, month = map(int, period.split('-'))
                days_in_month = monthrange(year, month)[1]
            except:
                days_in_month = 31
            
            schedule_data["employees"].extend(new_employees)
            for emp in new_employees:
                schedule_data["schedule"][emp["name"]] = [4] * days_in_month
            
            return self.save_schedule(period, schedule_data)

    # НОВЫЕ МЕТОДЫ ДЛЯ EXCEL ФАЙЛОВ
    
    def save_excel_to_server(self, file_path, created_by="Unknown"):
        """Сохранить Excel файл на сервер"""
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            try:
                with open(file_path, 'rb') as f:
                    file_data = base64.b64encode(f.read()).decode('utf-8')
                
                file_name = os.path.basename(file_path)
                response = self.network_manager.send_request({
                    'action': 'save_excel_file',
                    'file_name': file_name,
                    'file_data': file_data,
                    'created_by': created_by
                })
                return response and response.get('status') == 'success'
            except Exception as e:
                print(f"Error saving Excel to server: {e}")
                return False
        return False
    
    def get_excel_from_server(self, file_name, save_path, file_id=None):
        """Загрузить Excel файл с сервера"""
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            request_data = {'action': 'get_excel_file'}
            if file_id:
                request_data['file_id'] = file_id
            else:
                request_data['file_name'] = file_name
                
            response = self.network_manager.send_request(request_data)
            
            if response and response.get('status') == 'success':
                try:
                    file_data = base64.b64decode(response.get('file_data'))
                    with open(save_path, 'wb') as f:
                        f.write(file_data)
                    return True
                except Exception as e:
                    print(f"Error downloading Excel from server: {e}")
                    return False
        return False
    
    def get_available_excel_files(self):
        """Получить список доступных Excel файлов на сервере"""
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({
                'action': 'get_excel_files_list'
            })
            if response and response.get('status') == 'success':
                return response.get('data', [])
        return []
    
    def delete_excel_file(self, file_id):
        """Удалить Excel файл с сервера"""
        if self.use_network and self.network_manager and self.network_manager.is_connected:
            response = self.network_manager.send_request({
                'action': 'delete_excel_file',
                'file_id': file_id
            })
            return response and response.get('status') == 'success'
        return False


class MonthWidget(QWidget):
    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.parent = parent
        self.db = db
        self.current_period = None
        self.initUI()
    
    def initUI(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # Верхняя панель с выбором месяца
        top_layout = QHBoxLayout()
        
        self.period_combo = QComboBox()
        self.period_combo.currentIndexChanged.connect(self.period_changed)
        top_layout.addWidget(QLabel("Месяц:"))
        top_layout.addWidget(self.period_combo)
        
        top_layout.addStretch()
        layout.addLayout(top_layout)
        
        # Таблица
        self.table = NoteTableWidget(self)
        self.setup_table(self.table)
        layout.addWidget(self.table)
    
    def setup_table(self, table):
        table.setSelectionMode(QTableWidget.ExtendedSelection)
        table.setContextMenuPolicy(Qt.CustomContextMenu)
        table.customContextMenuRequested.connect(self.show_context_menu)
        
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        # Устанавливаем МЕНЬШУЮ высоту строк
        table.verticalHeader().setDefaultSectionSize(20)  # Увеличено с 25 до 30
        
        # Увеличиваем шрифт для заголовков дней
        font = table.horizontalHeader().font()
        font.setBold(True)
        font.setPointSize(12)  # Увеличено с 10 до 12
        table.horizontalHeader().setFont(font)
        
        # Увеличиваем шрифт для вертикальных заголовков (ФИО)
        font = table.verticalHeader().font()
        font.setBold(True)
        font.setPointSize(12)  # Увеличено с 10 до 12
        table.verticalHeader().setFont(font)
    
    def load_periods(self):
        periods = self.db.get_periods()
        self.period_combo.clear()
        
        if periods:
            for period in periods:
                try:
                    year, month = map(int, period.split('-'))
                    month_name = f"{self.parent.month_names[month]} {year}"
                    self.period_combo.addItem(month_name, period)
                except:
                    self.period_combo.addItem(period)
    
    def period_changed(self, index):
        if index >= 0:
            period = self.period_combo.itemData(index)
            self.current_period = period
            self.load_data(period)
    
    def get_working_days_count(self, year, month):
        """Возвращает количество рабочих дней в месяце (без воскресений)"""
        days_in_month = monthrange(year, month)[1]
        working_days = 0
        
        for day in range(1, days_in_month + 1):
            if datetime(year, month, day).weekday() != 6:  # 6 = воскресенье
                working_days += 1
                
        return working_days
    
    def get_day_mapping(self, year, month):
        """Создает маппинг между номером столбца и фактическим днем месяца (исключая воскресенья)"""
        days_in_month = monthrange(year, month)[1]
        day_mapping = []
        
        for day in range(1, days_in_month + 1):
            if datetime(year, month, day).weekday() != 6:  # Пропускаем воскресенья
                day_mapping.append(day)
                
        return day_mapping
    
    def load_data(self, period):
        # Устанавливаем флаг загрузки, чтобы избежать рекурсивных обновлений
        self.table.is_loading = True
        
        self.table.clear()
        self.current_period = period
            
        schedule_data = self.db.load_schedule(period)
        
        if not schedule_data:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.table.is_loading = False
            return
        
        employees_in_period = schedule_data.get("employees", [])
        if not employees_in_period:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            self.table.is_loading = False
            return
        
        try:
            year, month = map(int, period.split('-'))
            working_days_count = self.get_working_days_count(year, month)
            day_mapping = self.get_day_mapping(year, month)
            days_in_month = monthrange(year, month)[1]
        except:
            days_in_month = 31
            working_days_count = days_in_month
            day_mapping = list(range(1, days_in_month + 1))
        
        # Создаем таблицу с правильным количеством столбцов (только рабочие дни + 4 столбца подсчета)
        self.table.setRowCount(len(employees_in_period))
        self.table.setColumnCount(working_days_count + 4)  # рабочие дни + общее количество смен, регистратура, колл-центр, часы
        
        # Устанавливаем имена сотрудников в вертикальные заголовки
        self.table.setVerticalHeaderLabels([emp["name"] for emp in employees_in_period])
        
        notes_data = schedule_data.get('notes', {})
        
        for row, emp in enumerate(employees_in_period):
            emp_name = emp["name"]
            emp_schedule = schedule_data['schedule'].get(emp_name, [4]*days_in_month)  
            emp_notes = notes_data.get(emp_name, {})
            
            for col, actual_day in enumerate(day_mapping):  # Используем маппинг для столбцов
                day_index = actual_day - 1  # Индекс в исходном расписании
                day_status = emp_schedule[day_index]
                icon, _, bg_color, _ = self.parent.status_mapping.get(day_status, ("", "Пусто", QColor(Qt.white), 0))  # По умолчанию пусто
                
                note_key = str(day_index)
                has_note = note_key in emp_notes
                note_data = emp_notes.get(note_key, {})
                
                item = NoteItem(icon, has_note, note_data)
                item.setTextAlignment(Qt.AlignCenter)
                item.setBackground(bg_color)
                item.setForeground(Qt.black)
                
                if has_note:
                    end_time = note_data.get('end_time', '20:00')
                    item.setText(end_time)
                    item.setData(Qt.UserRole, True)
                    item.setToolTip(f"Отработано: {note_data.get('worked_hours', 12)} часов")
                else:
                    item.setText("")  # Пустая строка для всех статусов
                
                self.table.setItem(row, col, item)
        
        # Снимаем флаг загрузки
        self.table.is_loading = False
        
        # Обновляем заголовки и счетчики
        self.update_counters()
        
        # СИЛЬНО УМЕНЬШЕННАЯ ШИРИНА СТОЛБЦОВ С ДНЯМИ
        fixed_day_width = 25  # Уменьшено с 50 до 25 пикселей
        
        # Устанавливаем фиксированную ширину для всех столбцов с днями
        for col in range(working_days_count):
            self.table.setColumnWidth(col, fixed_day_width)
        
        # Ширина для столбцов подсчета (расширенные) - НОВЫЙ ПОРЯДОК
        self.table.setColumnWidth(working_days_count, 70)      # Общее количество смен (увеличено с 50 до 70)
        self.table.setColumnWidth(working_days_count + 1, 70)  # Регистратура (увеличено с 50 до 70)
        self.table.setColumnWidth(working_days_count + 2, 70)  # Колл-центр (увеличено с 50 до 70)
        self.table.setColumnWidth(working_days_count + 3, 80)  # Часы (увеличено с 60 до 80)
        
        # Настраиваем поведение заголовков
        # Столбцы с днями и подсчетами фиксированные
        for col in range(working_days_count + 4):
            self.table.horizontalHeader().setSectionResizeMode(col, QHeaderView.Fixed)
        
        # Установка увеличенной высоты строк
        self.table.verticalHeader().setDefaultSectionSize(30)  # Увеличено с 25 до 30
        
        self.table.viewport().update()

    def hours_to_hours_minutes(self, total_hours):
        """Конвертирует дробное количество часов в часы и минуты"""
        hours = int(total_hours)
        minutes = int(round((total_hours - hours) * 60))
        
        # Корректировка, если минуты равны 60
        if minutes == 60:
            hours += 1
            minutes = 0
            
        return hours, minutes

    def update_counters(self):
        """Обновляет все счетчики в реальном времени для таблицы"""
        if not self.current_period or self.table.is_loading:
            return
        
        try:
            year, month = map(int, self.current_period.split('-'))
            working_days_count = self.get_working_days_count(year, month)
            day_mapping = self.get_day_mapping(year, month)
            days_in_month = monthrange(year, month)[1]
        except:
            working_days_count = self.table.columnCount() - 4
            day_mapping = list(range(1, working_days_count + 1))
            days_in_month = working_days_count
        
        # Подсчет работающих по дням
        working_counts = [0] * working_days_count
        call_center_counts = [0] * working_days_count
        registry_counts = [0] * working_days_count
        
        for row in range(self.table.rowCount()):
            total_shifts = 0  # Общее количество смен
            call_center_days = 0
            registry_days = 0
            total_hours = 0.0
            
            for col in range(working_days_count):  # Исправлено: только рабочие дни
                item = self.table.item(row, col)
                if item:
                    bg_color = item.background().color()
                    
                    # Подсчет для сотрудника
                    if bg_color == self.parent.status_mapping[0][2]:  # Колл-центр
                        call_center_days += 1
                        total_shifts += 1
                        working_counts[col] += 1
                        call_center_counts[col] += 1
                        
                        # Подсчет часов с учетом примечания
                        if isinstance(item, NoteItem) and item.has_note:
                            total_hours += item.note_data.get('worked_hours', 12)
                        else:
                            total_hours += self.parent.status_mapping[0][3]
                            
                    elif bg_color == self.parent.status_mapping[1][2]:  # Регистратура
                        registry_days += 1
                        total_shifts += 1
                        working_counts[col] += 1
                        registry_counts[col] += 1
                        
                        # Подсчет часов с учетом примечания
                        if isinstance(item, NoteItem) and item.has_note:
                            total_hours += item.note_data.get('worked_hours', 12)
                        else:
                            total_hours += self.parent.status_mapping[1][3]
            
            # Конвертируем часы в формат "часы, минуты"
            hours, minutes = self.hours_to_hours_minutes(total_hours)
            
            # НОВЫЙ ПОРЯДОК СТОЛБЦОВ: общее количество смен, регистратура, колл-центр, часы
            
            # Общее количество смен
            if working_days_count < self.table.columnCount():
                total_shifts_item = self.table.item(row, working_days_count)
                if not total_shifts_item:
                    total_shifts_item = QTableWidgetItem()
                    total_shifts_item.setTextAlignment(Qt.AlignCenter)
                    total_shifts_item.setFlags(total_shifts_item.flags() ^ Qt.ItemIsEditable)
                    total_shifts_item.setFont(QFont("Arial", 11, QFont.Bold))  # Увеличено с 10 до 11
                    self.table.setItem(row, working_days_count, total_shifts_item)
                total_shifts_item.setText(f"{total_shifts}")
            
            # Регистратура
            if working_days_count + 1 < self.table.columnCount():
                registry_item = self.table.item(row, working_days_count + 1)
                if not registry_item:
                    registry_item = QTableWidgetItem()
                    registry_item.setTextAlignment(Qt.AlignCenter)
                    registry_item.setFlags(registry_item.flags() ^ Qt.ItemIsEditable)
                    registry_item.setFont(QFont("Arial", 11, QFont.Bold))  # Увеличено с 10 до 11
                    self.table.setItem(row, working_days_count + 1, registry_item)
                registry_item.setText(f"{registry_days}")
            
            # Колл-центр
            if working_days_count + 2 < self.table.columnCount():
                call_center_item = self.table.item(row, working_days_count + 2)
                if not call_center_item:
                    call_center_item = QTableWidgetItem()
                    call_center_item.setTextAlignment(Qt.AlignCenter)
                    call_center_item.setFlags(call_center_item.flags() ^ Qt.ItemIsEditable)
                    call_center_item.setFont(QFont("Arial", 11, QFont.Bold))  # Увеличено с 10 до 11
                    self.table.setItem(row, working_days_count + 2, call_center_item)
                call_center_item.setText(f"{call_center_days}")
            
            # Часы
            if working_days_count + 3 < self.table.columnCount():
                hours_item = self.table.item(row, working_days_count + 3)
                if not hours_item:
                    hours_item = QTableWidgetItem()
                    hours_item.setTextAlignment(Qt.AlignCenter)
                    hours_item.setFlags(hours_item.flags() ^ Qt.ItemIsEditable)
                    hours_item.setFont(QFont("Arial", 10, QFont.Bold))  # Увеличено с 9 до 10
                    self.table.setItem(row, working_days_count + 3, hours_item)
                
                # Форматируем текст в одну строку
                hours_text = f"{hours}ч {minutes}м"
                hours_item.setText(hours_text)
        
        # Обновляем заголовки дней с актуальными счетчиками
        day_headers = []
        for col, actual_day in enumerate(day_mapping):
            try:
                day_of_week = (datetime(year, month, actual_day).weekday())
                day_name = self.parent.day_names[day_of_week]
                count = working_counts[col]
                day_headers.append(f"{actual_day}\n{day_name}\n({count})")
            except:
                day_headers.append(str(actual_day))
        
        # Заголовки для всех столбцов - НОВЫЙ ПОРЯДОК
        headers = day_headers + ["Смены", "Рег", "КЦ", "Часы"]
        self.table.setHorizontalHeaderLabels(headers)

    def show_context_menu(self, pos):
        table = self.table
        selected = table.selectedIndexes()
        if not selected:
            return
        
        menu = QMenu()
        
        copy_action = QAction("Копировать", self)
        copy_action.triggered.connect(self.copy_selected)
        menu.addAction(copy_action)
        
        paste_action = QAction("Вставить", self)
        paste_action.triggered.connect(self.paste_selected)
        menu.addAction(paste_action)
        
        menu.addSeparator()
        
        if len(selected) == 1:
            index = selected[0]
            if index.column() < table.columnCount() - 4:  # Исключаем столбцы подсчета
                item = table.item(index.row(), index.column())
                if item and item.background().color() in [self.parent.status_mapping[0][2], self.parent.status_mapping[1][2]]:  # Только для колл-центра и регистратуры
                    if isinstance(item, NoteItem) and item.has_note:
                        remove_note_action = QAction("Удалить примечание", self)
                        remove_note_action.triggered.connect(lambda: self.remove_note(index))
                        menu.addAction(remove_note_action)
                        menu.addSeparator()
                    
                    note_action = QAction("Примечание", self)
                    note_action.triggered.connect(lambda: self.add_note(index))
                    menu.addAction(note_action)
                    menu.addSeparator()
        
        for status, (icon, text, _, _) in self.parent.status_mapping.items():
            action = QAction(f"■ {text}", self)  # Используем квадратик вместо иконки
            action.triggered.connect(lambda _, s=status: self.update_selected_status(s))
            menu.addAction(action)
        
        menu.exec_(table.viewport().mapToGlobal(pos))
    
    def add_note(self, index):
        item = self.table.item(index.row(), index.column())
        if not item or item.background().color() not in [self.parent.status_mapping[0][2], self.parent.status_mapping[1][2]]:  # Только для колл-центра и регистратуры
            return
        
        dialog = NoteDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            end_time = dialog.get_end_time()
            
            end_hour = int(end_time.split(':')[0])
            minutes = int(end_time.split(':')[1])
            worked_hours = end_hour - 8 + (minutes / 60.0)
            
            if isinstance(item, NoteItem):
                item.has_note = True
                item.note_data = {
                    'end_time': end_time,
                    'worked_hours': round(worked_hours, 2)
                }
                item.setText(end_time)
                item.setData(Qt.UserRole, True)
                item.setToolTip(f"Отработано: {round(worked_hours, 2)} часов")
                self.table.viewport().update()
                self.update_counters()  # Обновляем счетчики после добавления примечания
    
    def remove_note(self, index):
        item = self.table.item(index.row(), index.column())
        if item and isinstance(item, NoteItem):
            item.has_note = False
            item.note_data = {}
            item.setText("")  # Пустая строка вместо символа
            item.setData(Qt.UserRole, False)
            item.setToolTip("")
            self.table.viewport().update()
            self.update_counters()  # Обновляем счетчики после удаления примечания
    
    def copy_selected(self):
        selected = self.table.selectedIndexes()
        if not selected:
            return
            
        rows = defaultdict(list)
        for index in selected:
            rows[index.row()].append(index.column())
        
        self.copied_data = []
        for row, cols in rows.items():
            row_data = []
            for col in cols:
                item = self.table.item(row, col)
                # Копируем только цвет фона
                bg_color = item.background().color() if item else QColor(Qt.white)
                status = next((k for k, v in self.parent.status_mapping.items() if v[2] == bg_color), 4)
                row_data.append(status)
            self.copied_data.append(row_data)
    
    def paste_selected(self):
        if not hasattr(self, 'copied_data') or not self.copied_data:
            return
            
        selected = self.table.selectedIndexes()
        if not selected:
            return
            
        min_row = min(index.row() for index in selected)
        min_col = min(index.column() for index in selected)
        
        for row_offset, row_data in enumerate(self.copied_data):
            for col_offset, status in enumerate(row_data):
                target_row = min_row + row_offset
                target_col = min_col + col_offset
                
                if target_row >= self.table.rowCount() or target_col >= self.table.columnCount() - 4:  # Исключаем столбцы подсчета
                    continue
                
                item = self.table.item(target_row, target_col)
                if item and target_col < self.table.columnCount() - 4:  # Исправлено: исключаем столбцы подсчета
                    icon, _, bg_color, _ = self.parent.status_mapping.get(status, ("", "Пусто", QColor(Qt.white), 0))
                    
                    old_has_note = False
                    old_note_data = {}
                    if isinstance(item, NoteItem):
                        old_has_note = item.has_note
                        old_note_data = item.note_data
                    
                    new_item = NoteItem("", old_has_note, old_note_data)  # Пустой текст
                    new_item.setTextAlignment(Qt.AlignCenter)
                    new_item.setBackground(bg_color)
                    new_item.setForeground(Qt.black)
                    
                    if old_has_note:
                        new_item.setText(old_note_data.get('end_time', '20:00'))
                        new_item.setData(Qt.UserRole, True)
                        new_item.setToolTip(f"Отработано: {old_note_data.get('worked_hours', 12)} часов")
                    else:
                        new_item.setText("")  # Пустая строка
                    
                    self.table.setItem(target_row, target_col, new_item)
        
        self.update_counters()  # Обновляем счетчики после вставки
    
    def update_selected_status(self, status):
        selected = self.table.selectedIndexes()
        if not selected:
            return
        
        icon, _, bg_color, _ = self.parent.status_mapping[status]
        
        for index in selected:
            if index.column() >= self.table.columnCount() - 4:  # Исключаем столбцы подсчета
                continue
            
            item = self.table.item(index.row(), index.column())
            if item:
                old_has_note = False
                old_note_data = {}
                if isinstance(item, NoteItem):
                    old_has_note = item.has_note
                    old_note_data = item.note_data
                
                new_item = NoteItem("", old_has_note, old_note_data)  # Пустой текст
                new_item.setTextAlignment(Qt.AlignCenter)
                new_item.setBackground(bg_color)
                new_item.setForeground(Qt.black)
                
                if old_has_note:
                    new_item.setText(old_note_data.get('end_time', '20:00'))
                    new_item.setData(Qt.UserRole, True)
                    new_item.setToolTip(f"Отработано: {old_note_data.get('worked_hours', 12)} часов")
                else:
                    new_item.setText("")  # Пустая строка
                
                self.table.setItem(index.row(), index.column(), new_item)
        
        self.update_counters()  # Обновляем счетчики после изменения статуса


class ScheduleApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.settings = QSettings("MyCompany", "WorkSchedule")
        
        # Загрузка сохраненных настроек подключения
        saved_use_network = self.settings.value("use_network", False, type=bool)
        saved_host = self.settings.value("server_host", "localhost")
        saved_port = self.settings.value("server_port", 8888, type=int)
        
        # Диалог подключения к серверу
        connection_dialog = ServerConnectionDialog(
            self, 
            current_host=saved_host,
            current_port=saved_port,
            current_network_mode=saved_use_network
        )
        
        if connection_dialog.exec_() == QDialog.Accepted:
            params = connection_dialog.get_connection_params()
            try:
                self.db = ScheduleManager(
                    use_network=params['use_network'],
                    server_host=params['host'],
                    server_port=params['port']
                )
                # Сохраняем настройки
                self.settings.setValue("use_network", params['use_network'])
                self.settings.setValue("server_host", params['host'])
                self.settings.setValue("server_port", params['port'])
            except Exception as e:
                QMessageBox.warning(self, "Ошибка подключения", 
                                  f"Не удалось подключиться к серверу: {str(e)}\nБудет использован локальный режим.")
                self.db = ScheduleManager(use_network=False)
        else:
            # Локальный режим по умолчанию
            self.db = ScheduleManager(use_network=False)
        
        self.export_folder = self.settings.value("export_folder", "")
        
        # Убраны символы, оставлены только цвета
        self.status_mapping = {
            0: ("", "Колл-центр", QColor("#77FF77"), 12),  # Убран символ ↓
            1: ("", "Регистратура", QColor("#7777FF"), 12),  # Убран символ ↑
            2: ("", "Не работает", QColor("#FF7777"), 0),  # Убран символ ✕
            3: ("", "Отпуск", QColor("#FFFF77"), 0),  # Убран символ −
            4: ("", "Пусто", QColor("#FFFFFF"), 0)
        }
        
        self.month_names = {
            1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
            5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
            9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
        }
        
        self.day_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
        
        self.initUI()
        self.init_shortcuts()
        self.init_menu()
        self.load_initial_data()
    
    def initUI(self):
        self.setWindowTitle("Рабочее расписание" + (" [СЕТЕВОЙ РЕЖИМ]" if self.db.use_network else " [ЛОКАЛЬНЫЙ РЕЖИМ]"))
        self.setGeometry(100, 100, 1600, 800)
        # self.setWindowIcon(QIcon('schedule.ico'))
        
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout()
        self.central_widget.setLayout(self.layout)
        
        # Верхняя панель с кнопками
        self.buttons_layout = QHBoxLayout()
        
        # Левая часть панели - основные кнопки
        left_buttons_layout = QHBoxLayout()
        
        # Кнопка Новый месяц
        self.new_month_btn = QPushButton("Новый месяц")
        self.new_month_btn.clicked.connect(self.select_month_and_year)
        self.new_month_btn.setFixedHeight(40)  # высота
        self.new_month_btn.setFixedWidth(120)  # ширина
        self.new_month_btn.setStyleSheet("""
            QPushButton {
                background: #06B6D4;
                color: #FFFFFF;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton:hover {
                background: #0891B2;
            }
            QPushButton:pressed {
                background-color: #0E7490;
            }
        """)
        left_buttons_layout.addWidget(self.new_month_btn)
        
        # Кнопка Новый сотрудник
        self.add_employee_btn = QPushButton("Новый сотрудник")
        self.add_employee_btn.clicked.connect(self.add_employee_dialog)
        self.add_employee_btn.setFixedHeight(30)  # высота
        self.add_employee_btn.setFixedWidth(140)  # ширина
        self.add_employee_btn.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton:hover {
                background-color: #696969;
            }
            QPushButton:pressed {
                background-color: #505050;
            }
        """)
        left_buttons_layout.addWidget(self.add_employee_btn)
        
        # Кнопка для добавления сотрудников в текущий период
        self.add_to_period_btn = QPushButton("Добавить в месяц")
        self.add_to_period_btn.clicked.connect(self.add_employees_to_period)
        self.add_to_period_btn.setFixedHeight(30)  # высота
        self.add_to_period_btn.setFixedWidth(140)  # ширина
        self.add_to_period_btn.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton:hover {
                background-color: #696969;
            }
            QPushButton:pressed {
                background-color: #505050;
            }
        """)
        left_buttons_layout.addWidget(self.add_to_period_btn)
        
        # Правая часть панели - кнопка Сохранить и другие
        right_buttons_layout = QHBoxLayout()
        
        # Кнопка Открыть документ
        self.open_doc_btn = QPushButton("Открыть документ")
        self.open_doc_btn.clicked.connect(self.open_document)
        self.open_doc_btn.setFixedHeight(30)  # высота
        self.open_doc_btn.setFixedWidth(140)  # ширина
        self.open_doc_btn.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton:hover {
                background-color: #696969;
            }
            QPushButton:pressed {
                background-color: #505050;
            }
        """)
        right_buttons_layout.addWidget(self.open_doc_btn)
        
        # Кнопка Выбрать папку
        self.select_folder_btn = QPushButton("Выбрать папку")
        self.select_folder_btn.clicked.connect(self.select_export_folder)
        self.select_folder_btn.setFixedHeight(30)  # высота
        self.select_folder_btn.setFixedWidth(120)  # ширина
        self.select_folder_btn.setStyleSheet("""
            QPushButton {
                background-color: #808080;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton:hover {
                background-color: #696969;
            }
            QPushButton:pressed {
                background-color: #505050;
            }
        """)
        right_buttons_layout.addWidget(self.select_folder_btn)
        
        # Кнопка Сохранить - большая, зеленая, в правом углу
        self.save_btn = QPushButton("СОХРАНИТЬ")
        self.save_btn.clicked.connect(self.save_data)
        self.save_btn.setFixedHeight(40)  # высота
        self.save_btn.setFixedWidth(120)  # ширина
        self.save_btn.setStyleSheet("""
            QPushButton {
                background: #28A745;
                color: #FFFFFF;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 400;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:pressed {
                background-color: #1E7E34;
            }
        """)
        right_buttons_layout.addWidget(self.save_btn)
        
        # Объединяем левую и правую части
        self.buttons_layout.addLayout(left_buttons_layout)
        self.buttons_layout.addStretch()  # Растягивающийся элемент между левой и правой частями
        self.buttons_layout.addLayout(right_buttons_layout)
        
        self.layout.addLayout(self.buttons_layout)
        
        # Создаем сплиттер для вертикального расположения месяцев
        self.splitter = QSplitter(Qt.Vertical)
        self.layout.addWidget(self.splitter)
        
        # Создаем два виджета месяцев
        self.month_widget1 = MonthWidget(self, self.db)
        self.month_widget2 = MonthWidget(self, self.db)
        
        # Добавляем виджеты в сплиттер
        self.splitter.addWidget(self.month_widget1)
        self.splitter.addWidget(self.month_widget2)
        
        # Устанавливаем равные размеры
        self.splitter.setSizes([400, 400])
        
        self.legend_layout = QHBoxLayout()
        for status, (icon, text, color, _) in self.status_mapping.items():
            legend_item = QHBoxLayout()
            # Убраны иконки из легенды, оставлены только цвета
            color_label = QLabel("■")
            color_label.setStyleSheet(f"font-size: 16px; color: {color.name()}; padding: 2px;")
            legend_item.addWidget(color_label)
            
            text_label = QLabel(text)
            text_label.setStyleSheet("padding: 2px; font-size: 11px;")  # Увеличен шрифт легенды
            legend_item.addWidget(text_label)
            legend_item.addSpacing(20)
            self.legend_layout.addLayout(legend_item)
        
        self.legend_layout.addStretch()
        self.layout.addLayout(self.legend_layout)
        
        self.statusBar().showMessage("Готово" + (" [Подключено к серверу]" if self.db.use_network else " [Локальный режим]"))
    
    def init_menu(self):
        """Инициализация меню"""
        menubar = self.menuBar()
        
        # Меню "Файл"
        file_menu = menubar.addMenu('Файл')
        
        # Действие для настройки подключения
        connection_action = QAction('Настройки подключения...', self)
        connection_action.triggered.connect(self.show_connection_settings)
        file_menu.addAction(connection_action)
        
        file_menu.addSeparator()
        
        # Новые пункты для работы с общими файлами
        if self.db.use_network:
            shared_files_action = QAction('Общие Excel файлы...', self)
            shared_files_action.triggered.connect(self.show_shared_files)
            file_menu.addAction(shared_files_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction('Выход', self)
        exit_action.setShortcut('Ctrl+Q')
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Меню "Справка"
        help_menu = menubar.addMenu('Справка')
        
        about_action = QAction('О программе', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
    
    def init_shortcuts(self):
        for status, (icon, text, _, _) in self.status_mapping.items():
            shortcut = QShortcut(QKeySequence(f"Ctrl+{status}"), self)
            shortcut.activated.connect(lambda s=status: self.update_selected_status(s))
    
    def load_initial_data(self):
        # Загружаем периоды в оба виджета
        self.month_widget1.load_periods()
        self.month_widget2.load_periods()
        
        # Устанавливаем разные периоды по умолчанию, если доступно
        periods = self.db.get_periods()
        if periods:
            # Первый виджет - самый старый период
            self.month_widget1.period_combo.setCurrentIndex(0)
            # Второй виджет - самый новый период (последний в списке)
            if len(periods) > 1:
                self.month_widget2.period_combo.setCurrentIndex(len(periods) - 1)
            else:
                self.month_widget2.period_combo.setCurrentIndex(0)
    
    def show_connection_settings(self):
        """Показать диалог настроек подключения"""
        dialog = ServerConnectionDialog(
            self,
            current_host=self.db.server_host,
            current_port=self.db.server_port,
            current_network_mode=self.db.use_network
        )
        
        if dialog.exec_() == QDialog.Accepted:
            params = dialog.get_connection_params()
            
            try:
                # Пытаемся переподключиться с новыми параметрами
                success = self.db.reconnect(
                    use_network=params['use_network'],
                    server_host=params['host'],
                    server_port=params['port']
                )
                
                if success:
                    # Сохраняем настройки
                    self.settings.setValue("use_network", params['use_network'])
                    self.settings.setValue("server_host", params['host'])
                    self.settings.setValue("server_port", params['port'])
                    
                    # Обновляем интерфейс
                    self.setWindowTitle("Рабочее расписание" + (" [СЕТЕВОЙ РЕЖИМ]" if self.db.use_network else " [ЛОКАЛЬНЫЙ РЕЖИМ]"))
                    self.statusBar().showMessage("Настройки подключения обновлены" + (" [Подключено к серверу]" if self.db.use_network else " [Локальный режим]"), 3000)
                    
                    # Перезагружаем данные
                    self.load_initial_data()
                    
                    QMessageBox.information(self, "Успех", "Настройки подключения успешно обновлены")
                else:
                    QMessageBox.warning(self, "Ошибка", "Не удалось подключиться к серверу с новыми параметрами")
                    
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при изменении настроек подключения: {str(e)}")
    
    def show_shared_files(self):
        """Показать диалог общих файлов"""
        if self.db.use_network:
            dialog = SharedFilesDialog(self.db, self)
            dialog.exec_()
        else:
            QMessageBox.information(
                self, 
                "Информация", 
                "Эта функция доступна только в сетевом режиме.\n"
                "Перейдите в настройки подключения для включения сетевого режима."
            )
    
    def show_about(self):
        """Показать информацию о программе"""
        about_text = """
        <h3>Рабочее расписание</h3>
        <p>Программа для управления рабочими расписаниями сотрудников.</p>
        <p>Возможности:</p>
        <ul>
            <li>Создание и редактирование расписаний</li>
            <li>Учет рабочих часов и смен</li>
            <li>Экспорт в Excel</li>
            <li>Сетевой режим для совместной работы</li>
        </ul>
        <p><b>Текущий режим:</b> {}</p>
        """.format("СЕТЕВОЙ" if self.db.use_network else "ЛОКАЛЬНЫЙ")
        
        QMessageBox.about(self, "О программе", about_text)
    
    def select_month_and_year(self):
        """Открывает диалог выбора месяца и года"""
        dialog = MonthSelectionDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            selected_period = dialog.get_selected_period()
            
            # Проверяем, существует ли уже такой период
            existing_periods = [self.month_widget1.period_combo.itemData(i) for i in range(self.month_widget1.period_combo.count())]
            if selected_period in existing_periods:
                QMessageBox.information(self, "Информация", "Этот месяц уже существует")
                return
            
            self.create_new_period(selected_period)

    def create_new_period(self, period):
        """Создает новый период с выбранными сотрудниками"""
        employees = self.db.load_employees()
        if not employees:
            QMessageBox.warning(self, "Ошибка", "Нет доступных сотрудников")
            return
            
        dialog = EmployeeSelectionDialog(employees, self)
        if dialog.exec_() == QDialog.Accepted:
            selected_names = dialog.get_selected_employees()
            if not selected_names:
                QMessageBox.warning(self, "Ошибка", "Не выбрано ни одного сотрудника")
                return
                
            if self.db.create_period(period, selected_names):
                QMessageBox.information(self, "Успех", "Месяц создан")
                # Обновляем комбо-боксы в обоих виджетах
                self.month_widget1.load_periods()
                self.month_widget2.load_periods()
                
                # Устанавливаем новый период в первый виджет
                try:
                    year, month = map(int, period.split('-'))
                    month_name = f"{self.month_names[month]} {year}"
                    index = self.month_widget1.period_combo.findText(month_name)
                    if index >= 0:
                        self.month_widget1.period_combo.setCurrentIndex(index)
                except:
                    pass
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось создать месяц")
    
    def add_employees_to_period(self):
        """Добавляет сотрудников в уже созданный месяц"""
        # Используем текущий период из активного виджета
        current_widget = self.get_active_month_widget()
        if not current_widget or not current_widget.current_period:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите месяц в одном из окон")
            return
        
        employees = self.db.load_employees()
        if not employees:
            QMessageBox.warning(self, "Ошибка", "Нет доступных сотрудников")
            return
        
        schedule_data = self.db.load_schedule(current_widget.current_period)
        if not schedule_data:
            QMessageBox.warning(self, "Ошибка", "Не удалось загрузить данные выбранного месяца")
            return
        
        current_employees = schedule_data.get("employees", [])
        
        dialog = AddEmployeeToPeriodDialog(employees, current_employees, self)
        if dialog.exec_() == QDialog.Accepted:
            selected_names = dialog.get_selected_employees()
            if not selected_names:
                return
            
            if self.db.add_employees_to_period(current_widget.current_period, selected_names):
                QMessageBox.information(self, "Успех", f"Добавлено сотрудников: {len(selected_names)}")
                # Перезагружаем данные
                current_widget.load_data(current_widget.current_period)
            else:
                QMessageBox.warning(self, "Ошибка", "Не удалось добавить сотрудников")
    
    def get_active_month_widget(self):
        """Определяет, какой виджет месяца активен (имеет фокус)"""
        if self.month_widget1.table.hasFocus():
            return self.month_widget1
        elif self.month_widget2.table.hasFocus():
            return self.month_widget2
        else:
            # Если ни один не имеет фокуса, возвращаем первый
            return self.month_widget1
    
    def save_data(self):
        """Объединенная функция сохранения данных и экспорта в Excel"""
        # Сохраняем оба месяца
        for widget in [self.month_widget1, self.month_widget2]:
            if widget.current_period:
                schedule_data = self.prepare_save_data(widget)
                if schedule_data:
                    if self.db.save_schedule(widget.current_period, schedule_data):
                        self.statusBar().showMessage(f"Сохранено: {widget.current_period}", 2000)
                    else:
                        self.statusBar().showMessage("Ошибка сохранения", 5000)
                        return
        
        # Затем экспортируем в Excel
        self.export_to_excel()
    
    def prepare_save_data(self, widget):
        schedule_data = self.db.load_schedule(widget.current_period)
        if not schedule_data:
            return None
            
        try:
            year, month = map(int, widget.current_period.split('-'))
            days_in_month = monthrange(year, month)[1]
            working_days_count = widget.get_working_days_count(year, month)
            day_mapping = widget.get_day_mapping(year, month)
        except:
            days_in_month = widget.table.columnCount() - 4
            working_days_count = days_in_month
            day_mapping = list(range(1, days_in_month + 1))
        
        schedule_data["schedule"] = {}
        schedule_data["notes"] = {}
        
        for row in range(widget.table.rowCount()):
            emp_name = widget.table.verticalHeaderItem(row).text()
            full_schedule = [4] * days_in_month
            notes = {}
            
            for col, actual_day in enumerate(day_mapping):
                day_index = actual_day - 1  # Индекс в исходном расписании
                item = widget.table.item(row, col)
                if item and isinstance(item, NoteItem):
                    bg_color = item.background().color()
                    if bg_color == QColor(Qt.white):
                        status = 4
                    else:
                        status = next((k for k, v in self.status_mapping.items() if v[2] == bg_color), 4)
                    full_schedule[day_index] = status
                    
                    if item.has_note:
                        notes[str(day_index)] = item.note_data
            
            schedule_data["schedule"][emp_name] = full_schedule
            if notes:
                schedule_data["notes"][emp_name] = notes
        
        return schedule_data
    
    def update_selected_status(self, status):
        """Обновляет статус для активного виджета"""
        active_widget = self.get_active_month_widget()
        if active_widget:
            active_widget.update_selected_status(status)
    
    def add_employee_dialog(self):
        name, ok = QInputDialog.getText(self, "Добавить сотрудника", "ФИО сотрудника:")
        if ok and name:
            name = name.strip()
            if not name:
                QMessageBox.warning(self, "Ошибка", "Имя сотрудника не может быть пустым")
                return
                
            if self.db.add_employee(name):
                QMessageBox.information(self, "Успех", "Сотрудник добавлен")
            else:
                QMessageBox.warning(self, "Ошибка", "Сотрудник с таким именем уже существует")
    
    def open_document(self):
        """Открывает сохраненную таблицу Excel"""
        if not self.export_folder:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите папку для экспорта")
            return
        
        file_path = os.path.join(self.export_folder, "Расписание_все_месяцы.xlsx")
        
        if not os.path.exists(file_path):
            QMessageBox.warning(self, "Ошибка", f"Файл не найден:\n{file_path}")
            return
        
        try:
            # Открываем файл с помощью стандартного приложения
            if sys.platform == "win32":
                os.startfile(file_path)
            elif sys.platform == "darwin":  # macOS
                os.system(f'open "{file_path}"')
            else:  # linux
                os.system(f'xdg-open "{file_path}"')
            
            self.statusBar().showMessage(f"Открыт файл: {file_path}", 3000)
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть файл:\n{str(e)}")
        
    def select_export_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        if folder:
            self.export_folder = folder
            self.settings.setValue("export_folder", folder)
            QMessageBox.information(self, "Успех", f"Папка для экспорта установлена: {folder}")
    
    def export_to_excel(self):
        try:
            # Создаем Excel файл
            wb = self.create_excel_workbook()
            
            if self.db.use_network:
                # СЕТЕВОЙ РЕЖИМ - сохраняем на сервер
                
                # Создаем имя файла с временной меткой
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                file_name = f"Расписание_{timestamp}.xlsx"
                
                # Сохраняем во временный файл
                temp_dir = tempfile.gettempdir()
                temp_path = os.path.join(temp_dir, file_name)
                wb.save(temp_path)
                
                # Получаем имя текущего пользователя
                current_user = getpass.getuser()
                
                # Загружаем на сервер
                if self.db.save_excel_to_server(temp_path, created_by=current_user):
                    # Также сохраняем локально если указана папка
                    if self.export_folder:
                        local_path = os.path.join(self.export_folder, file_name)
                        wb.save(local_path)
                    
                    QMessageBox.information(
                        self, 
                        "Успех", 
                        f"Файл сохранен на сервер:\n{file_name}\n\n"
                        f"Все пользователи имеют доступ к этому файлу через меню 'Общие Excel файлы'."
                    )
                else:
                    QMessageBox.warning(self, "Ошибка", "Не удалось сохранить файл на сервер")
                
                # Удаляем временный файл
                try:
                    os.unlink(temp_path)
                except:
                    pass
                
            else:
                # ЛОКАЛЬНЫЙ РЕЖИМ - обычное сохранение
                if not self.export_folder:
                    self.select_export_folder()
                    if not self.export_folder:
                        return

                file_name = "Расписание_все_месяцы.xlsx"
                file_path = os.path.join(self.export_folder, file_name)
                wb.save(file_path)
                QMessageBox.information(self, "Успех", f"Файл сохранен:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать файл:\n{str(e)}")

    def create_excel_workbook(self):
        """Создает и заполняет Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"

        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        status_styles = {
            0: PatternFill(start_color="7FFF7F", end_color="7FFF7F", fill_type="solid"),
            1: PatternFill(start_color="8080FF", end_color="8080FF", fill_type="solid"),
            2: PatternFill(start_color="FF7777", end_color="FF7777", fill_type="solid"),
            3: PatternFill(start_color="FFFF77", end_color="FFFF77", fill_type="solid"),
            4: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        }

        # Цвета для столбцов подсчета
        total_shifts_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Оранжевый
        registry_fill = PatternFill(start_color="8080FF", end_color="8080FF", fill_type="solid")  # Синий
        call_center_fill = PatternFill(start_color="7FFF7F", end_color="7FFF7F", fill_type="solid")  # Зеленый
        hours_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Золотой

        # ФИКСИРОВАННЫЕ ПОЗИЦИИ ДЛЯ РЕЗУЛЬТИРУЮЩИХ СТОЛБЦОВ
        # Определяем максимальное количество рабочих дней среди всех месяцев
        max_working_days = 0
        periods = self.db.get_periods()
        
        for period in periods:
            schedule_data = self.db.load_schedule(period)
            if schedule_data and schedule_data.get('schedule'):
                try:
                    year, month = map(int, period.split('-'))
                    working_days_count = self.month_widget1.get_working_days_count(year, month)
                    max_working_days = max(max_working_days, working_days_count)
                except:
                    working_days_count = len(next(iter(schedule_data['schedule'].values())))
                    max_working_days = max(max_working_days, working_days_count)
        
        # Если не удалось определить, устанавливаем разумный максимум
        if max_working_days == 0:
            max_working_days = 31

        # ФИКСИРОВАННЫЕ ПОЗИЦИИ СТОЛБЦОВ
        RESULT_COLUMNS_START = max_working_days + 2  # +1 для столбца "Сотрудник", +1 для отступа
        RESULT_COLUMNS = {
            "Смены": RESULT_COLUMNS_START,
            "Рег": RESULT_COLUMNS_START + 1,
            "КЦ": RESULT_COLUMNS_START + 2,
            "Часы": RESULT_COLUMNS_START + 3
        }

        start_row = 1

        for period in periods:
            schedule_data = self.db.load_schedule(period)
            if not schedule_data or not schedule_data.get('schedule'):
                continue

            try:
                year, month = map(int, period.split('-'))
                month_name = f"{self.month_names[month]} {year}"
                working_days_count = self.month_widget1.get_working_days_count(year, month)
                day_mapping = self.month_widget1.get_day_mapping(year, month)
                days_in_month = monthrange(year, month)[1]
            except:
                month_name = period
                working_days_count = len(next(iter(schedule_data['schedule'].values())))
                day_mapping = list(range(1, working_days_count + 1))
                days_in_month = working_days_count

            employees = schedule_data.get('employees', [])
            if not employees:
                continue

            notes_data = schedule_data.get('notes', {})

            working_counts = [0] * working_days_count
            for emp in employees:
                emp_schedule = schedule_data['schedule'].get(emp["name"], [2]*days_in_month)
                for col, actual_day in enumerate(day_mapping):
                    day_index = actual_day - 1
                    if emp_schedule[day_index] in [0, 1]:
                        working_counts[col] += 1

            # Заголовок месяца - объединяем все столбцы включая результирующие
            total_columns = max_working_days + 5  # Сотрудник + дни + 4 результирующих столбца
            ws.cell(row=start_row, column=1, value=month_name).font = Font(bold=True, size=12)
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_columns)
            start_row += 1

            # Заголовки столбцов
            # Столбец "Сотрудник"
            ws.cell(row=start_row, column=1, value="Сотрудник").fill = header_fill
            ws.cell(row=start_row, column=1).alignment = center_alignment
            ws.cell(row=start_row, column=1).font = bold_font
            ws.cell(row=start_row, column=1).border = thin_border

            # Заголовки дней
            for col, actual_day in enumerate(day_mapping):
                try:
                    day_of_week = (datetime(year, month, actual_day).weekday())
                    day_name = self.day_names[day_of_week]
                    count = working_counts[col]
                    header_text = f"{actual_day}\n{day_name}\n({count})"
                except:
                    header_text = str(actual_day)
                
                cell = ws.cell(row=start_row, column=col+2, value=header_text)
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border

            # Заголовки результирующих столбцов в ФИКСИРОВАННЫХ позициях
            ws.cell(row=start_row, column=RESULT_COLUMNS["Смены"], value="Смены").fill = total_shifts_fill
            ws.cell(row=start_row, column=RESULT_COLUMNS["Смены"]).alignment = center_alignment
            ws.cell(row=start_row, column=RESULT_COLUMNS["Смены"]).font = bold_font
            ws.cell(row=start_row, column=RESULT_COLUMNS["Смены"]).border = thin_border

            ws.cell(row=start_row, column=RESULT_COLUMNS["Рег"], value="Рег").fill = registry_fill
            ws.cell(row=start_row, column=RESULT_COLUMNS["Рег"]).alignment = center_alignment
            ws.cell(row=start_row, column=RESULT_COLUMNS["Рег"]).font = bold_font
            ws.cell(row=start_row, column=RESULT_COLUMNS["Рег"]).border = thin_border

            ws.cell(row=start_row, column=RESULT_COLUMNS["КЦ"], value="КЦ").fill = call_center_fill
            ws.cell(row=start_row, column=RESULT_COLUMNS["КЦ"]).alignment = center_alignment
            ws.cell(row=start_row, column=RESULT_COLUMNS["КЦ"]).font = bold_font
            ws.cell(row=start_row, column=RESULT_COLUMNS["КЦ"]).border = thin_border

            ws.cell(row=start_row, column=RESULT_COLUMNS["Часы"], value="Часы").fill = hours_fill
            ws.cell(row=start_row, column=RESULT_COLUMNS["Часы"]).alignment = center_alignment
            ws.cell(row=start_row, column=RESULT_COLUMNS["Часы"]).font = bold_font
            ws.cell(row=start_row, column=RESULT_COLUMNS["Часы"]).border = thin_border

            start_row += 1

            # Устанавливаем ширину столбцов
            ws.column_dimensions['A'].width = 30  # Столбец с ФИО
            
            # Столбцы с днями - фиксированная ширина
            for col in range(2, max_working_days + 2):
                ws.column_dimensions[get_column_letter(col)].width = 4
            
            # Столбцы с результатами - фиксированная ширина в фиксированных позициях
            ws.column_dimensions[get_column_letter(RESULT_COLUMNS["Смены"])].width = 8
            ws.column_dimensions[get_column_letter(RESULT_COLUMNS["Рег"])].width = 8
            ws.column_dimensions[get_column_letter(RESULT_COLUMNS["КЦ"])].width = 8
            ws.column_dimensions[get_column_letter(RESULT_COLUMNS["Часы"])].width = 8

            # Данные сотрудников
            for emp in employees:
                emp_name = emp["name"]
                ws.cell(row=start_row, column=1, value=emp_name).font = bold_font
                
                schedule = schedule_data['schedule'].get(emp["name"], [2]*days_in_month)
                emp_notes = notes_data.get(emp["name"], {})
                total_shifts = 0
                total_hours = 0.0
                call_center_days = 0
                registry_days = 0
                
                # Данные по дням
                for col, actual_day in enumerate(day_mapping):
                    day_index = actual_day - 1
                    status = schedule[day_index]
                    
                    # Подсчет для результирующих столбцов
                    if status == 0:  # Колл-центр
                        call_center_days += 1
                        total_shifts += 1
                        if str(day_index) in emp_notes:
                            total_hours += emp_notes[str(day_index)].get('worked_hours', 12)
                        else:
                            total_hours += self.status_mapping[0][3]
                    elif status == 1:  # Регистратура
                        registry_days += 1
                        total_shifts += 1
                        if str(day_index) in emp_notes:
                            total_hours += emp_notes[str(day_index)].get('worked_hours', 12)
                        else:
                            total_hours += self.status_mapping[1][3]
                    
                    # Заполняем ячейку дня
                    note_key = str(day_index)
                    cell = ws.cell(row=start_row, column=col+2)
                    
                    if note_key in emp_notes:
                        end_time = emp_notes[note_key].get('end_time', '20:00')
                        cell.value = end_time
                    else:
                        cell.value = ""
                    
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    if status in status_styles:
                        cell.fill = status_styles[status]
                
                # Результирующие данные в ФИКСИРОВАННЫХ столбцах
                hours, minutes = self.hours_to_hours_minutes(total_hours)
                hours_text = f"{hours}ч {minutes}м"
                
                # Смены
                cell = ws.cell(row=start_row, column=RESULT_COLUMNS["Смены"], value=total_shifts)
                cell.fill = total_shifts_fill
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border
                
                # Рег
                cell = ws.cell(row=start_row, column=RESULT_COLUMNS["Рег"], value=registry_days)
                cell.fill = registry_fill
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border
                
                # КЦ
                cell = ws.cell(row=start_row, column=RESULT_COLUMNS["КЦ"], value=call_center_days)
                cell.fill = call_center_fill
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border
                
                # Часы
                cell = ws.cell(row=start_row, column=RESULT_COLUMNS["Часы"], value=hours_text)
                cell.fill = hours_fill
                cell.alignment = center_alignment
                cell.font = bold_font
                cell.border = thin_border
                
                start_row += 1

            start_row += 2

        return wb

    def hours_to_hours_minutes(self, total_hours):
        """Конвертирует дробное количество часов в часы и минуты"""
        hours = int(total_hours)
        minutes = int(round((total_hours - hours) * 60))
        
        # Корректировка, если минуты равны 60
        if minutes == 60:
            hours += 1
            minutes = 0
            
        return hours, minutes


if __name__ == "__main__":
    app = QApplication(sys.argv)
    font = app.font()
    font.setFamily("Arial")
    font.setPointSize(10)
    app.setFont(font)
    window = ScheduleApp()
    window.showMaximized()
    sys.exit(app.exec_())